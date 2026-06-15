#!/usr/bin/env python3
"""Fuel Manager — minimal HTTP server with REST API + static file serving."""

import json
import os
import threading
from http.server import HTTPServer, SimpleHTTPRequestHandler
from datetime import datetime

PORT = 8599
DATA_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(DATA_DIR, "dati.json")

# Default data structure
DEFAULT_DATA = {
    "persone": [
        "Luca", "Marco", "Daniel De Leo", "Federico Vessio", "Rebecca Fato",
        "Alessio Paparella", "Matteo Sangalli", "Giulia Pierin", "Alice Tonoli",
        "Ionuts Purniki", "Leonardo Scimone", "David Karpik", "Filippo Sciarpa",
        "Ilenia Boccagno", "Giorgia Bennici", "Sofia Vessio", "Romina Zeza",
        "Sonia Bertani", "Dennis Molinari", "Edoardo Ballarini", "Alessia Mongiardo",
        "Kerlins Hernandez", "Manuel Di Lanna", "Noemi Pianta", "Riccardo Carissimi",
        "Emanuele Meloni", "Marzio Foti", "Tommaso Signori", "Emanuele Radice",
        "Yuri Comelli", "Andres Lo Monaco", "Simone Lenoci", "Stella Massironi",
        "Zoe Nastrino", "Serena Malusardi", "Naike Maldifassi", "Simone Carcano"
    ],
    "tragitti": [],
    "benzina": [],
    "restituzioni": [],
    "spese": []
}

data_lock = threading.Lock()


def load_data():
    if os.path.exists(DATA_FILE):
        with open(DATA_FILE, "r") as f:
            return json.load(f)
    return DEFAULT_DATA.copy()


def save_data(data):
    with data_lock:
        with open(DATA_FILE, "w") as f:
            json.dump(data, f, ensure_ascii=False, indent=2)


class Handler(SimpleHTTPRequestHandler):
    def __init__(self, *args, **kwargs):
        super().__init__(*args, directory=DATA_DIR, **kwargs)

    def do_GET(self):
        if self.path == "/api/data" or self.path == "/api/data/":
            data = load_data()
            self._json_response(data)
        else:
            super().do_GET()

    def do_POST(self):
        if self.path == "/api/data" or self.path == "/api/data/":
            length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(length)
            data = json.loads(body)
            save_data(data)
            self._json_response({"ok": True})
        else:
            self._json_response({"error": "not found"}, 404)

    def do_PUT(self):
        if self.path == "/api/data" or self.path == "/api/data/":
            length = int(self.headers.get("Content-Length", 0))
            body = self.rfile.read(length)
            data = json.loads(body)
            save_data(data)
            self._json_response({"ok": True})
        else:
            self._json_response({"error": "not found"}, 404)

    def _json_response(self, obj, code=200):
        payload = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(code)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(payload)))
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()
        self.wfile.write(payload)

    def do_OPTIONS(self):
        self.send_response(204)
        self.send_header("Access-Control-Allow-Origin", "*")
        self.send_header("Access-Control-Allow-Methods", "GET, POST, PUT, OPTIONS")
        self.send_header("Access-Control-Allow-Headers", "Content-Type")
        self.end_headers()

    def log_message(self, format, *args):
        # Quieter logging
        pass


if __name__ == "__main__":
    # Seed from xlsx if dati.json doesn't exist
    if not os.path.exists(DATA_FILE):
        print("dati.json not found — seeding from xlsx...")
        try:
            import openpyxl

            wb = openpyxl.load_workbook(
                "/home/hermes/.hermes/cache/documents/doc_59b758883ca8_Gestione Benzina.xlsx",
                data_only=True,
            )

            def cv(v):
                if v is None: return None
                if isinstance(v, datetime): return v.strftime("%Y-%m-%d")
                if isinstance(v, str): return v.strip()
                return v

            def sf(v):
                if v is None: return 0
                if isinstance(v, (int, float)): return float(v)
                try: return float(str(v).replace(",", ".").strip())
                except: return 0

            def fc(hdrs, kws):
                for i, h in enumerate(hdrs):
                    if h is None: continue
                    for kw in kws:
                        if kw.lower() in str(h).lower(): return i
                return -1

            # Tragitti
            ws = wb["Tragitti"]
            h = [cv(c.value) for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
            di = fc(h, ["data"]); dk = fc(h, ["km", "tragitto"]); dc = fc(h, ["consumo"])
            dco = fc(h, ["costo", "carburante"]); dp = fc(h, ["persone"]); da = fc(h, ["aggiuntivi"])
            tragitti = []
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if not row or row[di] is None: continue
                data = cv(row[di]); km = sf(row[dk]); cons = sf(row[dc]); costo = sf(row[dco])
                pers = sf(row[dp]) if row[dp] else 1.55; agg = sf(row[da])
                tot = (km / 100) * cons * costo + agg
                tragitti.append({
                    "data": data, "km": round(km, 2), "consumo": round(cons, 2),
                    "costoCarburante": round(costo, 3), "persone": round(pers, 2),
                    "costiAgg": round(agg, 2), "totale": round(tot, 2),
                    "totaleTesta": round(tot / pers if pers > 0 else tot, 2),
                    "coeffKm": round(tot / km if km > 0 else 0, 6),
                })

            # Benzina
            ws = wb["Benzina"]
            h = [cv(c.value) for c in list(ws.iter_rows(min_row=1, max_row=1))[0]]
            di = fc(h, ["data"]); ds = fc(h, ["spesa"]); dco = fc(h, ["costo"])
            dl = fc(h, ["litri"]); dm = fc(h, ["mese"])
            benzina = []
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
                if not row or row[di] is None: continue
                data = cv(row[di]); spesa = sf(row[ds]); costo = sf(row[dco])
                litri = sf(row[dl]) if row[dl] else (spesa / costo if costo > 0 else 0)
                mese = str(row[dm] or "")[:10] if row[dm] else ""
                benzina.append({
                    "data": data, "spesa": round(spesa, 2), "costoCarburante": round(costo, 3),
                    "litri": round(litri, 2), "mese": mese,
                })

            # Restituzioni
            ws = wb["Restituzioni"]
            rn = [cv(c.value) for c in list(ws.iter_rows(min_row=2, max_row=2))[0]]
            nomi_r = []; cir = {}
            for i in range(2, len(rn)):
                n = rn[i]
                if n and n != "0": nomi_r.append(n); cir[n] = i
            restituzioni = []
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
                if not row or row[0] is None: continue
                data = cv(row[0]); det = {}; tot = 0
                for nome, idx in cir.items():
                    v = sf(row[idx])
                    if v != 0: det[nome] = round(v, 2); tot += v
                if tot > 0: restituzioni.append({"data": data, "totale": round(tot, 2), "dettaglio": det})

            # Spese
            ws = wb["Spese"]
            sn = [cv(c.value) for c in list(ws.iter_rows(min_row=2, max_row=2))[0]]
            nomi_s = []; cis = {}
            for i in range(2, len(sn)):
                n = sn[i]
                if n and n != "0": nomi_s.append(n); cis[n] = i
            spese = []
            for row in ws.iter_rows(min_row=3, max_row=ws.max_row, values_only=True):
                if not row or row[0] is None: continue
                data = cv(row[0]); det = {}; tot = 0
                for nome, idx in cis.items():
                    v = sf(row[idx])
                    if v != 0: det[nome] = round(v, 2); tot += v
                if tot > 0: spese.append({"data": data, "totale": round(tot, 2), "dettaglio": det})

            tutte = sorted(set(nomi_r + nomi_s))
            data = {"persone": tutte, "tragitti": tragitti, "benzina": benzina,
                    "restituzioni": restituzioni, "spese": spese}
            save_data(data)
            print(f"Seeded: {len(tragitti)} tragitti, {len(benzina)} benzina, "
                  f"{len(restituzioni)} rest, {len(spese)} spese, {len(tutte)} persone")
        except Exception as e:
            print(f"Seed error: {e}")

    server = HTTPServer(("0.0.0.0", PORT), Handler)
    print(f"Fuel Manager running on http://0.0.0.0:{PORT}")
    server.serve_forever()
